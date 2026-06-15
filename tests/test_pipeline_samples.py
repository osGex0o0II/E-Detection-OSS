from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from e_detection.batch import iter_csv_files, run_batch_detection
from e_detection.pipeline import check_anomaly_in_file
from e_detection.settings import DEFAULT_CONFIG, split_config

SAMPLE_DIR = Path(r"D:\System Files\Desktop\2025-09-06_电力监控系统运行数据")
LOW_VOLTAGE_SAMPLE = (
    SAMPLE_DIR
    / "1栋配电房"
    / "BM层低压配电房"
    / "综合楼1TM1低压系统"
    / "综合楼1TM1变压器运行日报_2025_09_06.csv"
)

pytestmark = pytest.mark.skipif(
    not SAMPLE_DIR.exists(),
    reason="真实样本目录不存在，跳过样本回归测试",
)


def test_sample_directory_shape():
    files = iter_csv_files(SAMPLE_DIR)

    assert len(files) == 43
    assert sum(1 for path in files if "高压" in str(path)) == 8


def test_low_voltage_sample_can_be_loaded_and_normalized():
    thresholds, enabled_rules = split_config(DEFAULT_CONFIG)
    _anomalies, log_data, cleaned_df, extra_info = check_anomaly_in_file(
        str(LOW_VOLTAGE_SAMPLE),
        thresholds,
        enabled_rules,
    )

    assert cleaned_df is not None
    assert len(cleaned_df) == 24
    assert {"Uab", "Ubc", "Uca", "Ia", "Ib", "Ic"}.issubset(cleaned_df.columns)
    assert isinstance(log_data, dict)
    assert extra_info["is_offline"] is False


def test_sample_batch_smoke_without_writing_report():
    result = run_batch_detection(SAMPLE_DIR, config=DEFAULT_CONFIG, write_report=False)

    assert result.total_files == 43
    assert result.processed_files == 43
    assert result.skipped_files >= 8
    assert result.report_path is None


def test_sample_batch_writes_xlsx_report(tmp_path: Path):
    result = run_batch_detection(
        SAMPLE_DIR,
        output_dir=tmp_path,
        config=DEFAULT_CONFIG,
        write_report=True,
    )

    assert result.report_path is not None
    assert result.report_path.suffix == ".xlsx"
    workbook = load_workbook(result.report_path, read_only=True)
    assert "异常明细" in workbook.sheetnames
    assert "检测概览" in workbook.sheetnames
