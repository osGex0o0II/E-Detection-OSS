from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from e_detection.excel_report import enrich_anomalies, write_excel_report
from e_detection.settings import DEFAULT_CONFIG


def test_write_excel_report_creates_expected_sheets_and_highlights(tmp_path: Path):
    anomalies = pd.DataFrame(
        [
            {
                "来源文件": "demo.csv",
                "日期": "2025-09-06",
                "异常类型": "电压异常; 疑似PT接线异常",
                "异常详情": "Uab过低",
                "异常值": "Uab=320",
                "时间": "0时",
                "Uab": 320.0,
                "Ubc": 408.0,
                "Uca": 409.0,
                "Ia": 10.0,
                "Ib": 11.0,
                "Ic": 12.0,
            }
        ]
    )
    enriched = enrich_anomalies(anomalies, "1栋配电房", "1TM1", "demo/demo.csv")
    report_path = tmp_path / "report.xlsx"

    write_excel_report(
        report_path,
        enriched,
        summary={
            "input_dir": "input",
            "output_dir": str(tmp_path),
            "total_files": 1,
            "processed_files": 1,
            "normal_files": 0,
            "anomaly_files": 1,
            "anomaly_records": 1,
            "skipped_files": 0,
            "generated_at": "2026-06-15 00:00:00",
        },
        config=DEFAULT_CONFIG,
    )

    workbook = load_workbook(report_path)
    assert workbook.sheetnames == [
        "检测概览",
        "异常明细",
        "设备汇总",
        "异常分类统计",
        "传感器状态",
        "检测配置",
    ]

    detail_sheet = workbook["异常明细"]
    headers = [cell.value for cell in detail_sheet[1]]
    assert "严重等级" in headers
    assert "建议处置" in headers

    severity_col = headers.index("严重等级") + 1
    uab_col = headers.index("Uab") + 1
    assert detail_sheet.cell(row=2, column=severity_col).value == "高"
    assert detail_sheet.cell(row=2, column=uab_col).fill.fgColor.rgb == "00FFF2CC"


def test_enrich_anomalies_adds_location_columns():
    anomalies = pd.DataFrame([{"来源文件": "demo.csv", "异常类型": "温度异常"}])

    enriched = enrich_anomalies(anomalies, "建筑A", "1TM1", "配电房/demo.csv")

    assert list(enriched.columns[:3]) == ["建筑", "相对路径", "变压器"]
    assert enriched.loc[0, "建筑"] == "建筑A"
    assert enriched.loc[0, "变压器"] == "1TM1"


def test_write_excel_report_escapes_formula_like_text(tmp_path: Path):
    anomalies = pd.DataFrame(
        [
            {
                "来源文件": "=cmd|' /C calc'!A0.csv",
                "日期": "2025-09-06",
                "异常类型": "@SUM(1,1)",
                "异常详情": "+malicious",
                "异常值": "-10",
                "时间": "\t=HYPERLINK(\"https://example.com\")",
            }
        ]
    )
    report_path = tmp_path / "formula-safe.xlsx"

    write_excel_report(
        report_path,
        enrich_anomalies(anomalies, "=建筑", "+变压器", "-相对路径"),
        summary={
            "input_dir": "=input",
            "output_dir": "+output",
            "total_files": 1,
            "processed_files": 1,
            "normal_files": 0,
            "anomaly_files": 1,
            "anomaly_records": 1,
            "skipped_files": 0,
            "generated_at": "@now",
        },
        config=DEFAULT_CONFIG,
        messages=["=message"],
    )

    workbook = load_workbook(report_path, data_only=False)
    overview_sheet = workbook["检测概览"]
    detail_sheet = workbook["异常明细"]

    assert overview_sheet["B2"].value == "'=input"
    assert overview_sheet["B3"].value == "'+output"
    assert overview_sheet["B10"].value == "'@now"
    assert overview_sheet["A15"].value == "'=message"

    detail_values = [cell.value for cell in detail_sheet[2]]
    assert "'=建筑" in detail_values
    assert "'-相对路径" in detail_values
    assert "'+变压器" in detail_values
    assert "'=cmd|' /C calc'!A0.csv" in detail_values
    assert "'@SUM(1,1)" in detail_values
    assert "'+malicious" in detail_values
    assert "'-10" in detail_values
    assert "'\t=HYPERLINK(\"https://example.com\")" in detail_values

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f"
