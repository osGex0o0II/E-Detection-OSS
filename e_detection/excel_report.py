"""Excel report generation for E-Detection."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .report_model import ReportContext, affected_columns, build_report_context, enrich_anomalies

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
SEVERITY_FILLS = {
    "高": PatternFill("solid", fgColor="F4CCCC"),
    "中": PatternFill("solid", fgColor="FCE5CD"),
    "低": PatternFill("solid", fgColor="D9EAD3"),
}
PARAMETER_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def write_excel_report(
    report_path: str | Path,
    anomalies: pd.DataFrame,
    summary: dict[str, Any],
    config: dict[str, Any] | None = None,
    sensor_status_rows: list[dict[str, Any]] | None = None,
    skipped_details: list[dict[str, Any]] | None = None,
    messages: list[str] | None = None,
) -> Path:
    """Write a multi-sheet XLSX report with row and cell highlights."""
    context = build_report_context(
        anomalies=anomalies,
        summary=summary,
        config=config,
        sensor_status_rows=sensor_status_rows,
        skipped_details=skipped_details,
        messages=messages,
    )
    return write_excel_report_from_context(report_path, context)


def write_excel_report_from_context(report_path: str | Path, context: ReportContext) -> Path:
    """Write a multi-sheet XLSX report from an internal report context."""
    report_file = Path(report_path)
    report_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview_ws = workbook.active
    overview_ws.title = "检测概览"

    _write_overview(overview_ws, context)
    _write_dataframe_sheet(
        workbook.create_sheet("异常明细"),
        context.details,
        context.config,
        highlight_details=True,
    )
    _write_dataframe_sheet(
        workbook.create_sheet("设备汇总"),
        context.device_summary,
        context.config,
    )
    _write_dataframe_sheet(
        workbook.create_sheet("异常分类统计"),
        context.type_statistics,
        context.config,
    )
    _write_dataframe_sheet(
        workbook.create_sheet("传感器状态"),
        context.sensor_status,
        context.config,
    )
    _write_config_sheet(workbook.create_sheet("检测配置"), context.config, vars(context.run))

    workbook.save(report_file)
    return report_file


def _write_overview(worksheet: Worksheet, context: ReportContext) -> None:
    worksheet.append(["指标", "值"])
    for key, label in [
        ("input_dir", "输入目录"),
        ("output_dir", "输出目录"),
        ("total_files", "文件总数"),
        ("processed_files", "已处理文件"),
        ("normal_files", "正常文件"),
        ("anomaly_files", "异常文件"),
        ("anomaly_records", "异常记录"),
        ("skipped_files", "跳过文件"),
        ("generated_at", "生成时间"),
    ]:
        worksheet.append([label, getattr(context.run, key)])

    worksheet.append([])
    worksheet.append(["高风险设备 Top 10", "异常记录数"])
    for _, row in context.high_risk_devices.head(10).iterrows():
        worksheet.append([f"{row['建筑']} / {row['变压器']}", int(row["异常记录数"])])

    if context.messages:
        worksheet.append([])
        worksheet.append(["逐文件结果摘录", ""])
        for message in context.messages[:20]:
            worksheet.append([message, ""])

    _style_tabular_sheet(worksheet)


def _write_dataframe_sheet(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    config: dict[str, Any],
    highlight_details: bool = False,
) -> None:
    if dataframe is None or dataframe.empty:
        dataframe = pd.DataFrame(columns=list(dataframe.columns) if dataframe is not None else [])

    worksheet.append(list(dataframe.columns))
    for _, row in dataframe.iterrows():
        worksheet.append([_excel_value(row.get(column)) for column in dataframe.columns])

    _style_tabular_sheet(worksheet)

    if highlight_details and not dataframe.empty:
        column_index = {cell.value: cell.column for cell in worksheet[1]}
        for row_number, (_, row) in enumerate(dataframe.iterrows(), start=2):
            severity = str(row.get("严重等级", "中"))
            fill = SEVERITY_FILLS.get(severity)
            if fill:
                for cell in worksheet[row_number]:
                    cell.fill = fill
            for column in affected_columns(row, config):
                if column in column_index:
                    worksheet.cell(row=row_number, column=column_index[column]).fill = PARAMETER_FILL


def _write_config_sheet(
    worksheet: Worksheet,
    config: dict[str, Any],
    summary: dict[str, Any],
) -> None:
    worksheet.append(["配置项", "值"])
    for key, value in config.items():
        worksheet.append([key, value])
    worksheet.append([])
    worksheet.append(["运行信息", "值"])
    for key, value in summary.items():
        worksheet.append([key, value])
    _style_tabular_sheet(worksheet)


def _style_tabular_sheet(worksheet: Worksheet) -> None:
    if worksheet.max_row >= 1:
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    if worksheet.max_column > 1 and worksheet.max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 42))
        worksheet.column_dimensions[letter].width = width

    worksheet.sheet_view.showGridLines = False


def _excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    return value
